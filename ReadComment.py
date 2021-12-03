

import os
from openpyxl import load_workbook, Workbook
from openpyxl.comments import Comment

def retPos(tpl):
    ww =tpl[0]
    ord_s = ord(ww)
    return ord_s-72

def retMizo(a,b):
    ret ='ERR'

    if (a ==1) :
        if (b[0] == 'G'):
            ret ="GHG"
        elif (b[0] == 'T'):
            ret  ="THA"
        else:
            ret ="ERR"

    if (a ==3) :
        if (b[0] == 'G'):
            ret ="GHS"
        elif (b[0] == 'T'):
            ret  ="THW"

    if (a ==0):
        if(b[0:2] == 'VM'):
            ret ="VM"
        elif(b[0:2] == 'VH'):
            ret ="VH"
     
    return ret






def xlAddr(wadr):
    import re

    s = wadr
    result = re.sub(r"\D", "", s)
 #   print(result)
    ii =s.find(result)
 #   print(ii)

#    print(s[0:ii])
    return(s[0:ii],result)



wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = '説明一覧'

os.chdir('d:\\SampPy')
wdir = os.getcwd()
print(wdir)

wb = load_workbook('demoa.xlsx')
ws = wb.active


ws_new['B2'] = '説明内容'

ws_new['C2'] = 'セル番地'   
ws_new['D2']  = 'Barcode'
ws_new['E2']  = '何番目'  
ws_new['F2']  = '形状'
ws_new['G2']  = 'Wide' 
ws_new['H2']  = 'Len' 

ws_new.column_dimensions['B'].width = 40
row_count = ws_new.max_row
print(f'新規作成したシートの行数は{row_count}です。')


for row in ws.iter_rows(min_row=1):
    for cell in row:
     
        if cell.comment is None:
            continue
        row_count = row_count + 1

    
        workAddr =str(cell.coordinate)
        print(workAddr)
        sss =xlAddr(workAddr)
        print(sss)



        wstr =cell.comment.text.replace(cell.comment.author+':','')


        ws_new[f'B{row_count}'] = wstr
        ws_new[f'C{row_count}'] = cell.coordinate
        ws_new[f'D{row_count}'] = ws.cell(row =int(sss[1]),column =3).value
        ws_new[f'E{row_count}'] = retPos(sss)
        ws_new[f'F{row_count}']  =retMizo(ws.cell(row =int(sss[1]),column =4).value,
                                          ws.cell(row =int(sss[1]),column =31).value)

        ws_new[f'G{row_count}'] = (ws.cell(row =int(sss[1]),column =5).value)/10
        ws_new[f'H{row_count}'] = (ws.cell(row =int(sss[1]),column =(retPos(sss)+8)).value)/10
        








ws_new['D2'].comment = Comment('説明があったセル番号', 'hogehoge')
wb_new.save('説明一覧.xlsx')

wb = load_workbook('説明一覧.xlsx')
ws = wb.active


f = open('myfile0000000.txt', 'w', encoding='shift_jis')


flag =True
j= 3

while flag:

    if (ws.cell(column =2,row =j).value ==None):
        flag  =False


    A1  =(ws.cell(column =1,row =j).value)
    if (A1 ==1):
        A1 ="追加工要"
    else:
        A1 =""
    A2  =(ws.cell(column =2,row =j).value)
 
    A2  =A2.strip('Takeshi Miki')
    A3  =(ws.cell(column =3,row =j).value)
    A4  =(ws.cell(column =4,row =j).value)
    A5  =(ws.cell(column =5,row =j).value)
    A6  =(ws.cell(column =6,row =j).value)
    A7  =(ws.cell(column =7,row =j).value)
    A8  =(ws.cell(column =8,row =j).value)
    print(A1)
    print(A2)
    print(A3)
    print(A4)
    print(A5)
    print(A6)
    print(A7)
    print(A8)


    datalist =[ '\n',
                'TMG(0)\n', 
                'LMG(0)\n',
                'EST(135)\n',
                'FMT(1,84,50,0,0,1)\n',
                'ACL()\n',
                'CFL(1,20,5,1,2,2)\n',
                'DAT(1,'+ A4 +'-'+ str(A5) +')\n',
                'CFL(2,20,15,1,2,2)\n',
                'DAT(2,'+ A6 +'-'+str(A7)+'*'+str(A8)+')\n',
                'CFL(3,20,25,1,2,2)\n',
                'DAT(3,'+ A2+')\n',
                'CFL(4,20,35,1,2,2)\n',
                'DAT(4,'+ A1+')\n',
                'PRT(1,1,1)\n'
                '\n',
                '\n' 
                ]
    f.writelines(datalist)

    
    j =j+1
 
    if (ws.cell(column =2,row =j).value ==None):
        flag  =False

f.close()   